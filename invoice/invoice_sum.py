#!/usr/bin/env python
# -*- coding: utf-8 -*-

import itertools
from openpyxl import load_workbook
import os.path
from openpyxl import Workbook
import pandas as pd
import numpy as np
import easygui as eg


### LOAD THE FILE


infile = 'Factures indépendant.xlsx'
filepath = os.path.abspath(infile)
#file = os.path.join(filepath, infile)
#filepath = os.path.abspath(os.pardir)
#file = os.path.join(filepath, filename)
#print(file)

xl = load_workbook(filepath, data_only=True)
sheet = xl.worksheets[0]

res = []

#iterate by rows
for row in sheet.iter_rows(min_row=1, min_col=1, max_col=8):
    res.append([cell.value for cell in row])
    #for cell in row:
    #    print(cell.value, end=" ")
    #print() 
    
#iterate by columns
#for row in sheet.iter_cols(min_row=1, min_col=1, max_col=8):
#    res.append([cell.value for cell in row])
#    #for cell in row:
#    #    print(cell.value, end=" ")


## CLEAN EMPTY LINES, SAVE TO DF

res2 = []
for i in res:
  res2.append([x for x in i if (x is not None)])
  
res2 = [x for x in res2 if x!=[]]

headers = res2.pop(0)
df = pd.DataFrame(res2, columns=headers)


## ASK USER WHAT TO SEARCH FOR

#popup box with choices
msg = "Find Possible Invoice Combinations"
title = "Find Invoices"
fieldNames = ["Amount to Find","Client","Maximum # of Invoices", "Largest Invoice # to check, inclusive (blank=check all)"]
fieldValues = ["0","ALTO","4",""]  # initial values
fieldValues = eg.multenterbox(msg,title, fieldNames, fieldValues)

value = float(fieldValues[0].strip().replace(',','.'))
client = fieldValues[1]
max_search = int(fieldValues[2].strip())+1
max_invoice_n = False

if fieldValues[3] != "":
  max_invoice_n = int(fieldValues[3].strip())


#value = 900 + 109.3
#client = 'ALTO'
#max_search = 3
#value = 900
#value = 3208.15
#value = 463.54

# filter by client, otherwise too large
df = df.loc[df['Client'] == client]
df = df[["N° facture","Client","TTC","Mission"]]
# correct TTC to round to 2 decimals
df['TTC'] = [round(x,2) for x in df['TTC']]

# filter by max_invoice_n if it exists
if max_invoice_n:
  df = df.loc[df['N° facture'] <= max_invoice_n ]

### FIND POSSIBLE COMBINATIONS
a = df['TTC']
ans = []

for i in range(1,min(len(a),max_search)): 
  #print("iteration ", i)
  for s in itertools.combinations(enumerate(a),i): 
    #print("s ",s)
    if np.abs(np.sum(s,axis=0)[1] - value) < 0.01:
      #print("s ",s)
      ans.append(list(s))

#print(ans) # contains list of indices and values that sum to value


## PREPARE TO OUTPUT
# now use ans to extract info from original df and store in res
res = []


for i in range(len(ans)):
  s = "\nPossible combination "+ str(i+1)
  #print(s)
  res.append(s)
  l = ans[i]
  #print("l ", l)
  rows = [int(idx) for (idx,_) in l] 
  #print("rows ", rows)
  #+2 because indices start at 1 for the header
    #print("rows ",rows)
  temp = df.iloc[rows,:]
  #print(temp)
  res.append(temp.to_string(index=False))

to_output = '\n'.join('{}'.format(k) for k in res)
print(to_output)
#    \x for x in res)

output_text = str(len(ans)) + " Possible Combinations for a sum of " + str(value) + "\n"

eg.textbox(output_text,"Output",to_output)




# simple example using itertools
#numbers = [1, 2, 3, 7, 7, 9, 10]
#result = [seq for i in range(len(numbers), 0, -1) for seq in itertools.combinations(numbers, i) if sum(seq) == 10]
#print(result)


